import pandas
import shutil
import os
from tkinter import filedialog, Tk
import openpyxl
import warnings
from openpyxl.styles import PatternFill
import sys

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# # Función para seleccionar el archivo ExcelEntrada
# def seleccionar_archivo():
#     root = Tk()
#     root.withdraw()  # Ocultar la ventana principal de Tkinter
#     archivo_entrada = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
#     return archivo_entrada


# Procesar el archivo RELACION LISTAS DE MATERIALES MP
def procesar_relacion_listas_mp(archivo_relacion):
    df_relacion = pandas.read_excel(archivo_relacion, sheet_name="M- MARCO SEATTLE", header=1, engine='openpyxl')

    df_relacion = df_relacion.iloc[0:, [2, 3, 7]]  # Extraemos las columnas C, D y H
    df_relacion.columns = ['NUMERO', 'REV', 'DESCRIPCION']  # Renombramos las columnas

    df_relacion['NUMERO'] = df_relacion['NUMERO'].fillna('').astype(str).str.strip()
    df_relacion['REV'] = df_relacion['REV'].fillna('').astype(str).str.replace('NO INDICA', 'NOIND')

    return df_relacion


# Procesar el archivo ExcelEntradaOLD
def procesar_excel_entrada_old(archivo_entrada_old):
    # Leer el archivo Excel en un DataFrame
    df_entrada_old = pandas.read_excel(archivo_entrada_old, sheet_name=0, header=None)

    # Obtener el nombre del archivo y extraer los valores numéricos y de revisión
    nombre_archivo = os.path.basename(archivo_entrada_old)
    num_excel_entrada_old = nombre_archivo.split('-')[1][:5]  # Extraer los primeros 5 dígitos después de 'M-'
    rev_excel_entrada_old = nombre_archivo.split('rev')[1].split(' ')[0]  # Extraer la revisión

    # Muestra los valores extraídos
    print(f"NumExcelEntradaOld: {num_excel_entrada_old}")
    print(f"RevExcelEntradaOld: {rev_excel_entrada_old}")

    # Extraer el valor de la celda E1 (posicionada en la fila 1, columna 4) del archivo Excel
    num_ensamble_entrada_old = df_entrada_old.iloc[0, 4]
    num_ensamble_entrada_old = str(num_ensamble_entrada_old).strip()  # Aseguramos que sea string y sin espacios
    if not num_ensamble_entrada_old:
        num_ensamble_entrada_old = 'Valor no disponible'

    # Muestra el valor de NumEnsambleEntrada
    print(f"NumEnsambleEntradaOld: {num_ensamble_entrada_old}")

    # Imprimir el DataFrame completo
    print("\nContenido de df_entrada_old:")
    print(df_entrada_old)

    return num_excel_entrada_old, rev_excel_entrada_old, num_ensamble_entrada_old, df_entrada_old


# Crear tabla para cada HojaImpresion, extrayendo datos según las especificaciones
def crear_tablas_por_hoja(df_entrada_old):
    tablas = []
    hoja_actual = 0
    sp_encontrado = False
    index_sp = None
    filas_inicio = 2  # Comienza desde la fila 3 (índice 2)

    # Verificar si hay SP, Sp o sp entre las primeras 300 filas
    for index, value in enumerate(df_entrada_old.iloc[:300, 0]):
        if value in ['SP', 'Sp', 'sp']:
            sp_encontrado = True
            index_sp = index
            break

    if not sp_encontrado:
        # Si no se encuentra SP, Sp o sp, la primera tabla será todo desde la fila 3 hasta el primer SP encontrado
        tabla = df_entrada_old.iloc[filas_inicio:, [0, 1, 2, 3, 4]]  # Toda la tabla a partir de la fila 3
        tabla.columns = ['PARTNO', 'QUANTITY', 'DESCRIPTION', 'MATERIAL', 'SPECIFICATION']
        tablas.append(tabla)

        # Imprimir la tabla creada para su visualización
        print(f"\nTabla 1 (Toda la tabla desde la fila 3):")
        print(tabla)
        print(f"  Número de filas en la tabla: {tabla.shape[0]}")
        print(f"  Columnas de la tabla: {', '.join(tabla.columns)}")
        print(f"  Primeros 5 registros de la tabla:")
        print(tabla.head())
    else:
        # Si encontramos SP, Sp o sp, seguimos la lógica de antes
        hoja_actual = 0
        # La primera tabla, desde la fila 3 hasta el primer SP encontrado
        tabla = df_entrada_old.iloc[filas_inicio:index_sp, [0, 1, 2, 3, 4]]
        tabla.columns = ['PARTNO', 'QUANTITY', 'DESCRIPTION', 'MATERIAL', 'SPECIFICATION']
        tablas.append(tabla)

        # Imprimir la primera tabla
        print(f"\nTabla 1 (Desde fila 3 hasta primer SP):")
        print(tabla)
        print(f"  Número de filas en la tabla: {tabla.shape[0]}")
        print(f"  Columnas de la tabla: {', '.join(tabla.columns)}")
        print(f"  Primeros 5 registros de la tabla:")
        print(tabla.head())

        hoja_actual = index_sp + 1  # Después del primer SP, empezamos la siguiente tabla

        # Luego, iteramos sobre los SP, Sp o sp para las tablas intermedias
        while hoja_actual < len(df_entrada_old):
            # Buscar el siguiente SP en el rango restante
            index_sp_siguiente = None
            for index, value in enumerate(df_entrada_old.iloc[hoja_actual:300, 0]):
                if value in ['SP', 'Sp', 'sp']:
                    index_sp_siguiente = index + hoja_actual
                    break

            if index_sp_siguiente is not None:
                # Tabla intermedia desde el último SP encontrado hasta el siguiente SP
                tabla = df_entrada_old.iloc[hoja_actual:index_sp_siguiente, [0, 1, 2, 3, 4]]
                tabla.columns = ['PARTNO', 'QUANTITY', 'DESCRIPTION', 'MATERIAL', 'SPECIFICATION']
                tablas.append(tabla)

                # Imprimir la tabla creada para su visualización
                print(f"\nTabla {len(tablas)} (Desde el último SP hasta el siguiente):")
                print(tabla)
                print(f"  Número de filas en la tabla: {tabla.shape[0]}")
                print(f"  Columnas de la tabla: {', '.join(tabla.columns)}")
                print(f"  Primeros 5 registros de la tabla:")
                print(tabla.head())

                hoja_actual = index_sp_siguiente + 1  # Continuar después del SP
            else:
                break  # Si no encontramos más SP, salimos del ciclo

        # La última tabla es desde el último SP hasta el final
        tabla_final = df_entrada_old.iloc[hoja_actual:, [0, 1, 2, 3, 4]]
        tabla_final.columns = ['PARTNO', 'QUANTITY', 'DESCRIPTION', 'MATERIAL', 'SPECIFICATION']
        tablas.append(tabla_final)

        # Imprimir la última tabla
        print(f"\nÚltima tabla (Desde el último SP hasta el final):")
        print(tabla_final)
        print(f"  Número de filas en la tabla: {tabla_final.shape[0]}")
        print(f"  Columnas de la tabla: {', '.join(tabla_final.columns)}")
        print(f"  Primeros 5 registros de la tabla:")
        print(tabla_final.head())

    return tablas


# Buscar fila en la TablaRelacionListasMP
def buscar_fila_en_relacion(df_relacion, num_excel_entrada):
    df_relacion['NUMERO'] = df_relacion['NUMERO'].str.strip()
    num_excel_entrada = str(num_excel_entrada).strip()
    filtro = df_relacion[df_relacion['NUMERO'] == num_excel_entrada]
    if not filtro.empty:
        fila = filtro.iloc[0]
        return fila['NUMERO'], fila['REV'], fila['DESCRIPCION']
    else:
        print(f"No se encontró el valor {num_excel_entrada} en la columna 'NUMERO'.")
        return None, None, None

from openpyxl.styles import Font, PatternFill, Alignment, Border

def copiar_estilos(origen, destino):
    # Copiar la fuente (fuente incluye el estilo, tamaño, color, etc.)
    if origen.font:
        destino.font = Font(name=origen.font.name,
                            size=origen.font.size,
                            bold=origen.font.bold,
                            italic=origen.font.italic,
                            underline=origen.font.underline,
                            color=origen.font.color)

    # Copiar el color de fondo
    if origen.fill:
        destino.fill = PatternFill(start_color=origen.fill.start_color,
                                   end_color=origen.fill.end_color,
                                   fill_type=origen.fill.fill_type)

    # Copiar la alineación
    if origen.alignment:
        destino.alignment = Alignment(horizontal=origen.alignment.horizontal,
                                      vertical=origen.alignment.vertical,
                                      text_rotation=origen.alignment.text_rotation,
                                      wrap_text=origen.alignment.wrap_text,
                                      shrink_to_fit=origen.alignment.shrink_to_fit)

    # Copiar los bordes
    if origen.border:
        destino.border = Border(left=origen.border.left,
                                right=origen.border.right,
                                top=origen.border.top,
                                bottom=origen.border.bottom,
                                diagonal=origen.border.diagonal,
                                diagonal_direction=origen.border.diagonal_direction,
                                outline=origen.border.outline,
                                vertical=origen.border.vertical,
                                horizontal=origen.border.horizontal)


def editar_archivo_entrada_excel(archivo_entrada, df_entrada, num_excel_relacion, rev_excel_relacion,
                                 descripcion_excel_relacion, tablas):
    # Cargar el archivo de Excel
    libro = openpyxl.load_workbook(archivo_entrada)
    rev_excel_entrada = archivo_entrada.split('rev')[1].split('.')[0]  # Obtener la parte después de 'rev'

    for hoja_actual, tabla in enumerate(tablas, start=1):
        hoja = libro.active

        # Definir la fila de inicio para cada hoja
        if hoja_actual == 1:
            start_row = 1  # La primera hoja empieza en la fila 1
        else:
            start_row = 54 + (hoja_actual - 2) * 50  # A partir de la segunda hoja, sumamos 50 filas por hoja

        # Ahora, llenamos las celdas de cada tabla en la hoja correspondiente
        for idx, row in tabla.iterrows():
            destino_fila = start_row + idx  # Ajustamos la fila de destino para cada ítem de la tabla
            # Aquí llenamos las celdas y copiamos los estilos, etc.

        # Cajetín superior
        hoja.cell(row=start_row + 1, column=2, value=descripcion_excel_relacion)  # DescripcionExcelRelacion
        hoja.cell(row=start_row + 1, column=8, value=df_entrada.iloc[0, 4])  # NumEnsambleEntrada

        # Imprimir lo que se está llenando en el cajetín superior
        print(f"\nLlenando cajetín superior en la hoja {hoja_actual}:")
        print(f"  DescripcionExcelRelacion: {descripcion_excel_relacion}")
        print(f"  NumEnsambleEntrada: {df_entrada.iloc[0, 4]}")
        print(f"  RevEnsambleEntrada: {rev_excel_entrada}")

        # Cajetín inferior
        hoja.cell(row=start_row + 48, column=8, value=f"M-{num_excel_relacion}")  # M-NumExcelRelacion
        if rev_excel_relacion == rev_excel_entrada:  # Compara con el rev del ExcelEntrada
            hoja.cell(row=start_row + 48, column=10, value=rev_excel_relacion)
            print(f"  REV coincide con el valor en ExcelEntrada.")
        else:
            hoja.cell(row=start_row + 48, column=10, value=rev_excel_relacion)
            print(f"  REV no coincide con el valor en ExcelEntrada, pero no se resalta en amarillo.")

        hoja.cell(row=start_row + 48, column=12, value=len(tablas))  # Total de hojas
        hoja.cell(row=start_row + 47, column=12, value=hoja_actual)  # HojaActual

        # Imprimir cajetín inferior
        print(f"  Total de hojas: {len(tablas)}")
        print(f"  Hoja actual: {hoja_actual}")

        # Llenado de ítems de listas
        print(f"\nLlenando ítems de listas en la hoja {hoja_actual}:")

        # Ajuste de filas de destino dentro de la tabla
        for idx, row in tabla.iterrows():
            destino_fila = start_row + 1 + idx  # Ajustar la fila de destino para cada ítem de la tabla
            # Partimos del principio para cada celda que escribiremos
            for col, field in [(2, 'PARTNO'), (3, 'QUANTITY'), (4, 'DESCRIPTION'), (7, 'MATERIAL'),
                               (8, 'SPECIFICATION')]:
                cell_destino = hoja.cell(row=destino_fila, column=col)
                value = row[field]

                # Verificar si la celda está fusionada
                merged_cell = None
                for merged_range in hoja.merged_cells.ranges:
                    if cell_destino.coordinate in merged_range:
                        merged_cell = merged_range
                        break

                if merged_cell is None:
                    # Si no está fusionada, escribir directamente
                    cell_destino.value = value
                    # Copiar estilos de la celda correspondiente
                    origen = hoja.cell(row=destino_fila, column=col)
                    copiar_estilos(origen, cell_destino)
                else:
                    # Si está fusionada, escribir solo en la celda principal de la fusión
                    main_cell = merged_cell.bounds[0], merged_cell.bounds[1]  # Top-left
                    cell_destino = hoja.cell(row=main_cell[0], column=main_cell[1])
                    cell_destino.value = value
                    # Copiar estilos de la celda correspondiente
                    origen = hoja.cell(row=main_cell[0], column=main_cell[1])
                    copiar_estilos(origen, cell_destino)
                    print(f"Cell {destino_fila}, column {col} is merged, writing to main cell.")

            # Imprimir cada ítem que se está llenando
            print(f"  Fila {destino_fila}:")
            print(f"    PARTNO: {row['PARTNO']}")
            print(f"    QUANTITY: {row['QUANTITY']}")
            print(f"    DESCRIPTION: {row['DESCRIPTION']}")
            print(f"    MATERIAL: {row['MATERIAL']}")
            print(f"    SPECIFICATION: {row['SPECIFICATION']}")

    # Guardar el archivo editado
    libro.save(archivo_entrada)
    print(f"\nArchivo editado: {archivo_entrada}")


import time
import subprocess
import os

def proceso_principal():
    # Selección de archivo de entrada
    # archivo_entrada = seleccionar_archivo()

    # Obtener el archivo de entrada desde los argumentos de la línea de comandos
    archivo_entrada = sys.argv[1]  # El archivo es el primer argumento

    print(f"Recibiendo el archivo: {archivo_entrada}")

    archivo_entrada_old = archivo_entrada.replace(".xlsx", " OLD.xlsx")

    archivo_relacion = 'D:/03. PROFESSIONAL EXP/02. MARCO PERUANA/UIPath+Python/RELACION LDM MARCO SEATTLE.xlsx'
    df_relacion = procesar_relacion_listas_mp(archivo_relacion)

    num_excel_entrada_old, rev_excel_entrada_old, num_ensamble_entrada_old, df_entrada_old = procesar_excel_entrada_old(archivo_entrada_old)

    tablas = crear_tablas_por_hoja(df_entrada_old)

    num_excel_relacion, rev_excel_relacion, descripcion_excel_relacion = buscar_fila_en_relacion(df_relacion,
                                                                                                 num_excel_entrada_old)

    if num_excel_relacion is not None:
        editar_archivo_entrada_excel(archivo_entrada, df_entrada_old, num_excel_relacion, rev_excel_relacion, descripcion_excel_relacion, tablas)

        # Esperamos 1 segundo antes de abrir los archivos
        time.sleep(1)

        # Obtener la carpeta donde está el archivo de entrada
        carpeta_entrada = os.path.dirname(archivo_entrada)

        # Buscar el único archivo PDF en la carpeta
        archivos_pdf = [f for f in os.listdir(carpeta_entrada) if f.endswith('.pdf')]

        if archivos_pdf:
            # Si encontramos un PDF, tomar el primer archivo (ya que debe ser el único)
            archivo_pdf = os.path.join(carpeta_entrada, archivos_pdf[0])

            # Abrir el archivo PDF con Microsoft Edge
            subprocess.run(['start', 'msedge', archivo_pdf], shell=True)

        # Abrir el archivo Excel
        subprocess.run(['start', 'excel', archivo_entrada], shell=True)

# Ejecutar el proceso
if __name__ == "__main__":
    proceso_principal()
