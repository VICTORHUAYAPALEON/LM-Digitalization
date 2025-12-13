import subprocess
import os

import openpyxl
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook
import shutil
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle


# Función para abrir el archivo Excel mediante una ventana emergente
def abrir_archivo():
    Tk().withdraw()  # Evitar que se abra una ventana vacía de Tkinter
    archivo = filedialog.askopenfilename(title="Seleccionar archivo Excel Entrada",
                                         filetypes=[("Archivos Excel", "*.xlsx")])
    return archivo


# Función para obtener información del archivo ExcelEntrada
def procesar_excel_entrada(archivo_entrada):
    # Leer el nombre del archivo
    archivo_nombre = os.path.basename(archivo_entrada)
    num_excel_entrada = archivo_nombre.split(" ")[0][2:]  # Extraer el número
    rev_excel_entrada = archivo_nombre.split("rev")[1].split(".")[0]  # Extraer la revisión

    # Mostrar en consola
    print(f"NumExcelEntrada: {num_excel_entrada}")
    print(f"RevExcelEntrada: {rev_excel_entrada}")

    # Leer el archivo Excel
    df_entrada = pd.read_excel(archivo_entrada, sheet_name=None)  # Lee todas las hojas

    # Verificar si la hoja "Hoja1" o "Sheet1" existe y ajustar el nombre de la hoja
    if 'Hoja1' in df_entrada:
        num_ensamble_entrada = df_entrada['Hoja1'].iloc[0, 4]  # Leer valor de la celda E1 en "Hoja1"
    elif 'Sheet1' in df_entrada:
        num_ensamble_entrada = df_entrada['Sheet1'].iloc[0, 4]  # Leer valor de la celda E1 en "Sheet1"
    else:
        raise ValueError("No se encontró 'Hoja1' ni 'Sheet1' en el archivo Excel")

    # Intentar convertir el valor
    try:
        num_ensamble_entrada = int(num_ensamble_entrada)
    except ValueError:
        num_ensamble_entrada = str(num_ensamble_entrada)

    # Verificar si está vacío o nulo
    if pd.isna(num_ensamble_entrada):
        print("La celda E1 está vacía o contiene un valor nulo.")
    else:
        print(f"NumEnsambleEntrada: {num_ensamble_entrada}")
    return num_excel_entrada, rev_excel_entrada, num_ensamble_entrada



def procesar_relacion_listas_mp():
    # Cargar el archivo Excel
    archivo = r'D:\03. PROFESSIONAL EXP\02. MARCO PERUANA\UIPath+Python\RELACION LDM MARCO SEATTLE.xlsx'

    # Leer la hoja "M- MARCO SEATTLE", comenzando desde la fila 3 (índice 2 en pandas)
    df_relacion = pd.read_excel(archivo, sheet_name='M- MARCO SEATTLE', header=1, skiprows=[0])

    # Seleccionar las columnas C, D y H, que corresponden a los índices 2, 3 y 7
    df_relacion = df_relacion.iloc[:, [2, 3, 7]]  # C, D y H
    df_relacion.columns = ['NUMERO', 'REV', 'DESCRIPCION']  # Renombramos las columnas

    # Asegurarnos de que los valores en la columna 'REV' sean cadenas de texto y reemplazar "NO INDICA" por "NOIND"
    df_relacion['REV'] = df_relacion['REV'].apply(lambda x: str(x).strip() if x != 'NO INDICA' else 'NOIND')

    # Asegurarnos de que los valores en la columna 'NUMERO' sean cadenas de texto
    df_relacion['NUMERO'] = df_relacion['NUMERO'].apply(lambda x: str(x).strip())



    # Mostrar el DataFrame resultante para verificar que todo esté bien
    print(df_relacion.head())

    return df_relacion


# Función para contar páginas (separadores)
def contar_paginas(df):
    contador_paginas = 1  # Comienza en la primera página
    print("DataFrame original:")
    print(df.head())  # Imprimir las primeras filas del DataFrame para ver cómo está estructurado
    for i, row in df.iterrows():
        # Acceso por índice, ya que no hay nombre de columna
        valor = str(row[0]).strip().lower()  # Normalizamos el valor (a string) y quitamos espacios
        print(f"Fila {i}: '{row[0]}' -> '{valor}'")  # Mostrar cómo se procesa cada fila
        if valor == 'sp':  # Verificar si es un separador de página
            print(f"Se detectó 'SP' en la fila {i}")
            contador_paginas += 1
    return contador_paginas


# Función para crear el archivo de salida
def crear_archivo_salida(archivo_entrada, num_excel_entrada, rev_excel_entrada, contador_paginas):
    # Obtener la ruta del archivo de entrada
    ruta_entrada = os.path.dirname(archivo_entrada)  # Obtener la carpeta donde está el archivo de entrada
    nombre_archivo_entrada = os.path.basename(archivo_entrada)  # Obtener el nombre del archivo de entrada

    # Crear el nuevo nombre para el archivo de entrada con el sufijo OLD
    nuevo_nombre_entrada = f"{nombre_archivo_entrada.replace('.xlsx', '')} OLD.xlsx"
    ruta_nuevo_entrada = os.path.join(ruta_entrada, nuevo_nombre_entrada)

    # Renombrar el archivo de entrada añadiendo "OLD"
    if os.path.exists(archivo_entrada):
        os.rename(archivo_entrada, ruta_nuevo_entrada)
        print(f"Archivo de entrada renombrado a: {ruta_nuevo_entrada}")
    else:
        print(f"El archivo de entrada no existe en la ruta: {archivo_entrada}")
        return

    # Ruta del archivo de formato (ajústalo según tu ruta real)
    formato_archivo = r'D:\03. PROFESSIONAL EXP\02. MARCO PERUANA\UIPath+Python\PY-MP-005 PostprocesamientoLM\M-XXXXX FORMATO.xlsx'  # Ruta estática al archivo de formato

    # Nombre del nuevo archivo de salida basado en el nombre del archivo de entrada (sin el sufijo OLD)
    nuevo_archivo_salida = f"{nombre_archivo_entrada.replace(' OLD', '')}"  # Mantiene el nombre original, sin el "OLD"
    ruta_nuevo_archivo = os.path.join(ruta_entrada, nuevo_archivo_salida)  # Guardar en la misma carpeta que el archivo de entrada

    # Crear copia del archivo de formato en la misma ruta del archivo de entrada
    shutil.copy(formato_archivo, ruta_nuevo_archivo)
    print(f"Archivo de formato copiado y renombrado a: {ruta_nuevo_archivo}")

    # Cargar el nuevo archivo
    wb_salida = load_workbook(ruta_nuevo_archivo)
    ws_salida = wb_salida.active

    # Lógica para pegar los datos en el archivo de salida según el contador de páginas
    for i in range(1, contador_paginas + 1):
        # Dependiendo del contador de páginas, determinar el rango de celdas a copiar
        if i > 1:  # Si hay más de una página, dejar una fila en blanco
            fila_inicio = (i - 1) * 50 + 1  # Calcula la fila de inicio para cada página
            ws_salida.append([''])  # Insertar una fila vacía

    # Guardar el archivo de salida con las modificaciones
    wb_salida.save(ruta_nuevo_archivo)
    print(f"Archivo de salida guardado como: {ruta_nuevo_archivo}")


def copiar_rango_con_formato(ws_origen, ws_salida, rango_origen, fila_destino):
    for row in ws_origen[rango_origen]:
        for cell in row:
            # Obtener la celda de destino
            destino_cell = ws_salida.cell(row=fila_destino, column=cell.column)

            # Copiar el valor de la celda
            destino_cell.value = cell.value

            # Copiar solo los atributos del estilo de manera más controlada
            if cell.has_style:
                destino_cell.font = cell.font
                destino_cell.fill = cell.fill
                destino_cell.border = cell.border
                destino_cell.alignment = cell.alignment

            fila_destino += 1

def ajustar_rango_y_borrar(ws_salida, contador_paginas):
    # Mostrar el valor del contador de páginas
    print(f"Contador de páginas: {contador_paginas}")

    # Determinar el rango máximo según el contador de hojas
    if contador_paginas == 1:
        fila_max = 50
    elif contador_paginas == 2:
        fila_max = 100
    elif contador_paginas == 3:
        fila_max = 150
    elif contador_paginas == 4:
        fila_max = 200
    else:
        fila_max = 200  # Definir un límite por defecto para más de 4 páginas

    # Mostrar el rango máximo para verificar
    print(f"Rango máximo para eliminar hasta la fila: {fila_max}")

    # Eliminar todas las filas después del rango máximo
    if fila_max < ws_salida.max_row:  # Si hay más filas que el rango máximo
        ws_salida.delete_rows(fila_max + 1, ws_salida.max_row - fila_max)

    print(f"Se ajustaron las filas hasta el rango: {fila_max}")

def procesar_archivo_entrada(archivo_entrada, contador_paginas):
    # Cargar el archivo de entrada
    wb_entrada = openpyxl.load_workbook(archivo_entrada)
    ws_entrada = wb_entrada.active

    # Ajustar el rango y borrar filas superiores según el contador de páginas
    ajustar_rango_y_borrar(ws_entrada, contador_paginas)

    # Guardar el archivo modificado con las filas eliminadas
    wb_entrada.save(archivo_entrada)  # Sobrescribe el archivo de entrada
    print("Archivo de entrada modificado y guardado.")



# Función principal que coordina el flujo de trabajo
def main():
    # Paso 1: Solicitar archivo ExcelEntrada
    archivo_entrada = abrir_archivo()

    # Paso 2: Procesar ExcelEntrada
    num_excel_entrada, rev_excel_entrada, num_ensamble_entrada = procesar_excel_entrada(archivo_entrada)

    # Paso 3: Procesar RELACION LISTAS DE MATERIALES MP
    df_relacion = procesar_relacion_listas_mp()

    # Paso 4: Seleccionar fila basada en NumExcelEntrada

    # Aquí convertimos num_excel_entrada en cadena (string) para asegurar que la comparación sea válida
    num_excel_entrada = str(num_excel_entrada).strip()  # Convertir a string y eliminar espacios

    fila_relacion = df_relacion[df_relacion['NUMERO'] == num_excel_entrada].iloc[0]
    num_excel_relacion = fila_relacion['NUMERO']
    rev_excel_relacion = fila_relacion['REV']
    descripcion_excel_relacion = fila_relacion['DESCRIPCION']
    print(f"NumExcelRelacion: {num_excel_relacion}")
    print(f"RevExcelRelacion: {rev_excel_relacion}")
    print(f"DescripcionExcelRelacion: {descripcion_excel_relacion}")

    # Paso 5: Contar páginas

    # Leer solo las primeras 200 filas del archivo de Excel en un DataFrame
    df_relacion = pd.read_excel(archivo_entrada, header=None, nrows=200)
    contador_paginas = contar_paginas(df_relacion)
    print(f"Número total de páginas: {contador_paginas}")

    # Paso 6: Crear archivo de salida
    crear_archivo_salida(archivo_entrada, num_excel_relacion, rev_excel_relacion, contador_paginas)

    # Paso 7: Crear archivo de salida
    procesar_archivo_entrada(archivo_entrada, contador_paginas)

    # Aquí pasa el archivo_entrada al otro script "PY-MP-005-02 WriteDataTemplate.py"
    print(f"Pasando el archivo {archivo_entrada} al script PY-MP-005-02 WriteDataTemplate.py...")

    # Ejecutamos el otro script, pasando el archivo_entrada como argumento
    subprocess.run(['python', 'PY-MP-005-02 WriteDataTemplate.py', archivo_entrada])

    print(f"El archivo {archivo_entrada} se ha pasado al script PY-MP-005-02 WriteDataTemplate.py.")



if __name__ == "__main__":
    main()
